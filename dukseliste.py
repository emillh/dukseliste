import pandas as pd
import openpyxl
from openpyxl import Workbook

class Student:
    def __init__(self, name:str):
        self.name = name
        self.sweep = 0
        self.trash = 0
        self.table = 0
        self.total = 0
        self.last_week = False
    
    def add_sweep(self):
        self.sweep += 1
        self._add_total()
    
    def add_trash(self):
        self.trash += 1
        self._add_total()

    def add_table(self):
        self.table += 1
        self._add_total()

    def _add_total(self):
        self.total += 1
        self.last_week = True
    
    def get_sweep(self):
        return self.sweep
    
    def get_trash(self):
        return self.trash
    
    def get_table(self):
        return self.table
    
    def get_last_week(self):
        return self.last_week

def generate_week(students: list) -> list:
    week = [None] * 6

    students.sort(key=lambda x: x.sweep + x.total)
    for student in students:
        if week[0] is None and student not in week and student.get_last_week() is False:
            week[0] = student
        elif week[1] is None and student not in week and student.get_last_week() is False:
            week[1] = student
        
        if week[0] is not None and week[1] is not None and student not in week and student.get_last_week() is False:
            if student.get_sweep() < week[0].get_sweep():
                week[0] = student
            elif student.get_sweep() < week[1].get_sweep():
                week[1] = student
        
    
    students.sort(key=lambda x: x.trash + x.total)
    for student in students:    
        if week[2] is None and student not in week and student.get_last_week() is False:
            week[2] = student
        elif week[3] is None and student not in week and student.get_last_week() is False:
            week[3] = student
        
        if week[2] is not None and week[3] is not None and student not in week and student.get_last_week() is False:
            if student.get_trash() < week[2].get_trash():
                week[2] = student
            elif student.get_trash() < week[3].get_trash():
                week[3] = student
    
    students.sort(key=lambda x: x.table + x.total)
    for student in students:
        if week[4] is None and student not in week and student.get_last_week() is False:
            week[4] = student
        elif week[5] is None and student not in week and student.get_last_week() is False:
            week[5] = student
        
        if week[4] is not None and week[5] is not None and student not in week and student.get_last_week() is False:
            if student.get_table() < week[4].get_table():
                week[4] = student
            elif student.get_table() < week[5].get_table():
                week[5] = student
    
    for student in students:
        student.last_week = False

    week[0].add_sweep()
    week[1].add_sweep()
    week[2].add_trash()
    week[3].add_trash()
    week[4].add_table()
    week[5].add_table()

    return week

def generate_sheet(students: list, weeks: int) -> None:
    new_workbook = Workbook()

    new_sheet = new_workbook.active

    new_sheet.cell(row=1, column=1, value="Uge")
    new_sheet.cell(row=1, column=2, value="Pligt 1 - Feje")
    new_sheet.cell(row=1, column=3, value="Pligt 2 - Feje")
    new_sheet.cell(row=1, column=4, value="Pligt 3 - Affald")
    new_sheet.cell(row=1, column=5, value="Pligt 4 - Affald")
    new_sheet.cell(row=1, column=6, value="Pligt 5 - Tørre borde")
    new_sheet.cell(row=1, column=7, value="Pligt 6 - Tørre borde")

    for week in range(1, weeks + 1):
        chores = generate_week(students)
        for col, chore in enumerate(chores, start=2):
            new_sheet.cell(row=week + 1, column=col, value=chore.name)

    new_workbook.save("dukseliste.xlsx")

if __name__ == "__main__":
    workbook = openpyxl.load_workbook("elever.xlsx")

    first_sheet = workbook[workbook.sheetnames[0]]

    students = []
    for row in first_sheet.iter_rows(min_row=1, max_row=first_sheet.max_row, min_col=1, max_col=2, values_only=True):
        name, _ = row
        new_student = Student(name)
        students.append(new_student)

    school_weeks = 40

    generate_sheet(students, school_weeks)
